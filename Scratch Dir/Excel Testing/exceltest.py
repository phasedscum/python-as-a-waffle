from openpyxl import workbook, load_workbook

xl = load_workbook(filename='./ipaddr.xlsx', data_only=True)
xs = xl.active
col_rang = xl['Sheet1']
cell_rang = xs['A1':'A2']
#print(col_rang['A'].value)
print ('Getting Data From :', xl.get_sheet_names())
for row in xs.iter_rows():
    for cell in row:
        ipaddr = cell.value
        #print(ipaddr)
        print(file=open(ipaddr + ".txt", "a"))